import openpyxl
from yagsend import Tinson

# Run this program for execution

def Surya():
    book = openpyxl.load_workbook('Excel.xlsx')
    sheet = book.get_sheet_by_name('main')
    row_count = sheet.max_row

    # Fetching senders details from the sheet

    for i in range(2, 3):
        I = str(sheet.cell(row = i, column = 1).value)
        me = str(sheet.cell(row = i, column = 2).value)
        password = str(sheet.cell(row = i, column = 3).value)
        phone = "+" + str(sheet.cell(row = i, column = 4).value)
        link = str(sheet.cell(row = i, column = 5).value)
        altemail = str(sheet.cell(row = i, column = 6).value)
        Tinson(I, me, password, altemail, link, phone)

if __name__ == '__main__':
    Surya()