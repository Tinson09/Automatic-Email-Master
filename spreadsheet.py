import openpyxl

SUB_SHEET_NAME = "main"
SHEET_PARAM_NAME_ROW = 1
FIRST_COLUMN_INDEX = 1
EMAIL_KEY = "email"


class Row:
    def __init__(self):
        self.row_data = dict()
        self.key_set = []

    def __init__(self, row_data):
        self.row_data = dict(row_data)
        self.key_set = self.row_data.keys()

    def get_value(self, key):
        return self.row_data[key]

    def set_value(self, key, value):
        self.row_data[key] = value
        self.key_set = self.row_data.keys()

    def get_key_set(self):
        return self.key_set

    def get_email(self):
        if EMAIL_KEY in self.row_data:
            return self.row_data[EMAIL_KEY]
        else:
            return None


class Data:
    def __init__(self):
        self.data = []

    def __init__(self, data):
        self.data = list(data)

    def get_row(self, index):
        return self.data[index]

    def set_row(self, index, row):
        self.data.insert(index, row)

    def add_row(self, row):
        self.data.append(row)

    def get_count(self):
        return len(self.data)


class SpreadSheet:
    def __init__(self, file_location):
        self.file_location = str(file_location)
        book = openpyxl.load_workbook(self.file_location)
        sheet = book.get_sheet_by_name(SUB_SHEET_NAME)
        self.row_count = sheet.max_row
        self.column_count = sheet.max_column
        self.field_key = {}
        self.data = Data()
        for index in range(FIRST_COLUMN_INDEX, self.column_count):
            key = str(sheet.cell(SHEET_PARAM_NAME_ROW, index).value)
            self.field_key[index] = key

        for i in range(SHEET_PARAM_NAME_ROW + 1, self.row_count):
            row = Row()
            for j in range(FIRST_COLUMN_INDEX, self.column_count):
                value = str(sheet.cell(i, j).value)
                row.set_value(self.field_key[j], value)
            self.data.add_row(row)

    def get_data(self):
        return self.data
