import yagmail
from draft import Neha, Amal
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from dateformat import reformat_date
import sys, os
from multiprocessing import Process

def blockPrint():
    sys.stdout = open(os.devnull, 'w')

def enablePrint():
    sys.stdout = sys.__stdout__

def cout(str):
    enablePrint()
    print(str)
    blockPrint()

def Bombay(yag, HR_email, cc, I, company, html):
    try:
        yag.send(to = {HR_email:I}, cc = cc, subject = company + ": Placement Cell, Govt. Model Engineering College", contents = html)
        cout("Done " + company)
    except:
        cout("Sending Failed for "+company)

def Tinson(I, me, password, altemail, link, phone):
    ok = True
    jobs = []
    try:
        yag = yagmail.SMTP(me, password)
        cout(I + " Authenticated")
    except:
        ok = False
        cout("Connection Error")
    if ok:
        book = openpyxl.load_workbook('Excel.xlsx')
        sheet = book.get_sheet_by_name(I)
        row_count = sheet.max_row
        print "Number of rows : " + str(row_count)

        #Following section takes uniques data that should be included inside each mail from the sheet
        for i in range(2, row_count + 1):
            cc = []
            company = str(sheet.cell(row = i, column = 1).value)
            HR = str(sheet.cell(row = i, column = 2).value)
            gender = str((sheet.cell(row = i, column = 3).value).encode("utf8"))
            #gender = "Sir"
            HR_email = str(sheet.cell(row = i, column = 4).value)
            reciever = [HR_email]
            Appointment = sheet.cell(row = i, column = 5).value
            Appointment = reformat_date(str(Appointment))
            secondary = str(sheet.cell(row = i, column = 6).value)
            cc2 = str(sheet.cell(row = i, column = 7).value)
            draft = str(sheet.cell(row = i, column = 8).value)
            link = str(sheet.cell(row = i, column = 9).value)
            cc.append(secondary)
            if cc2 != 'None':
                cc.append(xMECian)
            #reciever = reciever+cc


            # Function Neha and Amal returns a draft, arguments are unique details that should be included in the draft
            # You can choose b/w 2 types of draft. However if you have only one draft Mark as 1 in the sheet and use only function Neha

            reciever.append(cc)
            if int(draft) == 1:
                text, html = Neha(I, HR, gender, company, phone, me, altemail, Appointment, link)
            elif int(draft) == 2:
                text, html = Amal(I, HR, gender, company, phone, me, altemail, Appointment, link)

 
            #p=Process(target=Bombay,args=[yag,HR_email,cc,I,company,html])
            #jobs.append(p)
            #p.start()



                    # Following sections sends the mail

            try:
                #Edit the following line to specify a subject
                yag.send(to = {HR_email:I}, cc = cc, subject = company + ": Placement Cell, Govt. Model Engineering College", contents = html)
                cout("Done " + company)
            except smtplib.SMTPAuthenticationError:
                cout("Sending Failed for " + company)  